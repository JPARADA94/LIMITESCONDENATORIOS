# LIMITES CONDENATORIOS.py
# Autor: Javier Parada
# Versión: 2.0 — Edición Profesional
# Entrada: Excel exportado desde SmartAssistance (formato ARCHIVO 2)
# Llave de análisis: COMPONENTE
# Salida: Excel con formato + CSV con límites de Precaución y Condenatorio

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import plotly.graph_objects as go
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# Configuración de página
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Límites Condenatorios — Análisis de Lubricación",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────
# CSS profesional
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Segoe UI', 'Inter', sans-serif; }

.main-header {
    background: linear-gradient(135deg, #1a3a5c 0%, #1d6fcf 100%);
    color: white;
    padding: 2rem 2.5rem 1.6rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
}
.main-header h1 { margin: 0; font-size: 1.9rem; font-weight: 700; letter-spacing: -0.4px; }
.main-header p  { margin: 0.4rem 0 0; font-size: 0.92rem; opacity: 0.85; }

.kpi-card {
    background: white;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    text-align: center;
    box-shadow: 0 1px 4px rgba(0,0,0,0.07);
}
.kpi-value { font-size: 1.8rem; font-weight: 700; color: #1a3a5c; line-height: 1.1; }
.kpi-label { font-size: 0.74rem; color: #6b7280; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }

.section-title {
    font-size: 1.1rem;
    font-weight: 600;
    color: #1a3a5c;
    border-left: 4px solid #1d6fcf;
    padding-left: 0.75rem;
    margin: 2rem 0 0.9rem;
}

.footer {
    text-align: center;
    font-size: 0.78rem;
    color: #9ca3af;
    margin-top: 3rem;
    padding-top: 1rem;
    border-top: 1px solid #e5e7eb;
}

.block-container { padding-top: 1.2rem; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Constantes: variables, unidades, iconos de categoría
# ─────────────────────────────────────────────────────────────
GUIDE_VARS = {
    "Desgaste": [
        "PLATA (AG) - 19", "ALUMINIO (AL) - 20", "CROMO (CR) - 24",
        "COBRE (CU) - 25", "HIERRO (FE) - 26", "ÍNDICE PQ (PQI) - 3",
        "NÍQUEL (NI) - 32", "PLOMO (PB) - 35", "ESTAÑO (SN) - 37", "TITANIO (TI) - 38",
    ],
    "Propiedades del lubricante": [
        "NÚMERO BÁSICO (BN) - 12", "VISCOSIDAD A 100 °C - 13",
        "NÚMERO ÁCIDO (AN) - 43", "OXIDACIÓN - 80", "NITRACIÓN - 82",
    ],
    "Contaminantes": [
        "CADMIO (CD) - 23", "POTASIO (K) - 27", "MANGANESO (MN) - 29",
        "SODIO (NA) - 31", "SILICIO (SI) - 36", "VANADIO (V) - 39",
        "HOLLÍN - 79", "AGUA (IR) - 74",
    ],
    "Aditivos": [
        "BORO (B) - 18", "BARIO (BA) - 21", "CALCIO (CA) - 22",
        "MAGNESIO (MG) - 28", "MOLIBDENO (MO) - 30", "FÓSFORO (P) - 34", "ZINC (ZN) - 40",
    ],
}

VAR_UNITS = {
    "PLATA (AG) - 19": "ppm",       "ALUMINIO (AL) - 20": "ppm",
    "CROMO (CR) - 24": "ppm",       "COBRE (CU) - 25": "ppm",
    "HIERRO (FE) - 26": "ppm",      "ÍNDICE PQ (PQI) - 3": "índice",
    "NÍQUEL (NI) - 32": "ppm",      "PLOMO (PB) - 35": "ppm",
    "ESTAÑO (SN) - 37": "ppm",      "TITANIO (TI) - 38": "ppm",
    "NÚMERO BÁSICO (BN) - 12": "mg KOH/g",
    "VISCOSIDAD A 100 °C - 13": "cSt",
    "NÚMERO ÁCIDO (AN) - 43": "mg KOH/g",
    "OXIDACIÓN - 80": "abs/cm",     "NITRACIÓN - 82": "abs/cm",
    "CADMIO (CD) - 23": "ppm",      "POTASIO (K) - 27": "ppm",
    "MANGANESO (MN) - 29": "ppm",   "SODIO (NA) - 31": "ppm",
    "SILICIO (SI) - 36": "ppm",     "VANADIO (V) - 39": "ppm",
    "HOLLÍN - 79": "%",             "AGUA (IR) - 74": "%",
    "BORO (B) - 18": "ppm",         "BARIO (BA) - 21": "ppm",
    "CALCIO (CA) - 22": "ppm",      "MAGNESIO (MG) - 28": "ppm",
    "MOLIBDENO (MO) - 30": "ppm",   "FÓSFORO (P) - 34": "ppm",
    "ZINC (ZN) - 40": "ppm",
}

CAT_ICONS = {
    "Desgaste": "⚙️",
    "Propiedades del lubricante": "🧪",
    "Contaminantes": "⚠️",
    "Aditivos": "🔬",
}

# ─────────────────────────────────────────────────────────────
# Encabezado principal
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🔧 Límites Condenatorios por Componente</h1>
    <p>Análisis estadístico de lubricación · Histórico SmartAssistance · v2.0</p>
</div>
""", unsafe_allow_html=True)

with st.expander("📖 ¿Qué hace esta herramienta y cómo usarla?", expanded=False):
    st.markdown("""
    ### ¿Para qué sirve?
    Esta herramienta te ayuda a **definir cuándo un componente está en problemas**, basándose en el
    historial real de tus análisis de aceite. En lugar de usar límites genéricos del fabricante,
    calcula límites **específicos para tu flota**, usando sus propios datos históricos.

    El resultado son dos alertas por variable (hierro, cobre, agua, etc.):
    - 🟡 **Límite de Precaución** — El valor está subiendo. Hay que poner el componente en vigilancia
      y planificar una inspección. No es urgente, pero no lo ignores.
    - 🔴 **Límite Condenatorio** — El valor está muy por encima de lo normal. El componente
      probablemente tiene un problema activo. Considera sacarlo de servicio o intervenir pronto.

    ---
    ### Paso a paso
    | Paso | Qué hacer | Por qué |
    |------|-----------|---------|
    | **1** | Carga uno o más Excel de SmartAssistance | La app combina todos los archivos automáticamente. No modifica los archivos. |
    | **2** | Aplica filtros si quieres (opcional) | Para calcular límites específicos por operación, tipo de equipo o lubricante. Si no filtras, usa todo el historial. |
    | **3** | Revisa el inventario | Verifica cuántas muestras tiene cada componente. Con pocas muestras, los límites son menos confiables. |
    | **4** | Elige el modo de cálculo | *Por componente*: límites distintos para cada equipo. *Mezcla*: un único límite compartido (útil cuando tienes pocos datos por equipo). |
    | **5** | Selecciona los componentes | Los que quieres analizar en esta sesión. |
    | **6** | Marca las variables | Las sustancias que te importa vigilar (hierro, agua, viscosidad, etc.). |
    | **7** | Presiona "Calcular límites" | La app calcula y muestra los resultados listos para exportar. |

    ---
    ### ¿Qué significan los colores del resultado?
    - 🟡 **Fondo amarillo** en la columna de Precaución → ese es tu umbral de alerta temprana.
    - 🔴 **Fondo rojo** en la columna Condenatorio → ese es tu umbral de acción inmediata.
    - ⬜ **Fondo gris** → no había suficientes muestras para calcular ese límite con confianza.

    > **Importante:** los límites calculados son tan buenos como el historial que los respalda.
    > Un componente con 5 muestras dará límites menos confiables que uno con 50.
    """)

# ─────────────────────────────────────────────────────────────
# Sidebar — Parámetros de cálculo
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Parámetros de cálculo")

    with st.expander("❓ ¿Para qué sirven estos parámetros?", expanded=False):
        st.caption(
            "Estos controles ajustan **cómo se calculan los límites**. "
            "Si no estás seguro, deja los valores por defecto — están calibrados para "
            "análisis de aceite estándar. Solo cámbialos si tienes un criterio técnico específico."
        )

    min_n = st.number_input(
        "Mínimo de datos válidos",
        min_value=2, value=3, step=1,
        help=(
            "Cantidad mínima de análisis que necesita un componente para calcularle límites. "
            "Con 1 o 2 muestras el resultado no sería estadísticamente confiable, "
            "así que la app lo marca como 'Insuficiente' y no calcula. "
            "Valor recomendado: 3 a 5."
        )
    )

    st.markdown("**Umbral de método**")
    n_switch = st.number_input(
        "Usar percentiles si n ≥",
        min_value=3, value=10, step=1,
        help=(
            "Cuando un componente tiene MUCHAS muestras (n ≥ este valor), la app usa percentiles "
            "(método más robusto y preciso). Cuando tiene POCAS muestras, usa promedio + desviación "
            "(método más conservador). "
            "Valor recomendado: 10."
        )
    )

    st.markdown("**Percentiles** *(historial largo, n ≥ umbral)*")
    st.caption(
        "P90 = el 90% de tus análisis históricos estuvieron por debajo de este valor. "
        "Es decir: si el hierro llega a P90, ya está en el 10% más alto que has visto. Eso es una señal."
    )
    p_prec  = st.slider("Percentil Precaución",    50, 99, 90, 1)
    p_alert = st.slider("Percentil Condenatorio",  50, 99, 95, 1)

    st.markdown("**Factores k** *(historial corto, n < umbral)*")
    st.caption(
        "Cuando hay pocos datos, el límite se calcula como: "
        "Promedio + k × Desviación. "
        "k=2 captura ~95% de los valores normales. k=3 captura ~99.7%. "
        "Valores mayores de k = límites más permisivos."
    )
    k_prec  = st.number_input("k Precaución  (μ + k·σ)",    min_value=0.0, value=2.0, step=0.5)
    k_alert = st.number_input("k Condenatorio (μ + k·σ)",  min_value=0.0, value=3.0, step=0.5)

    st.markdown("**Opciones avanzadas**")
    usar_iqr = st.checkbox(
        "Limpiar outliers (IQR) antes de calcular",
        value=False,
        help=(
            "Activa esto si tienes muestras con valores extremos sospechosos "
            "(errores de laboratorio, contaminación puntual, etc.) que inflarían los límites. "
            "La app descarta automáticamente los valores muy alejados del rango normal. "
            "Especialmente útil para la variable PQI."
        )
    )

    st.divider()
    st.markdown("**ℹ️ Aviso legal**")
    st.caption(
        "© 2026 · Javier Parada. Todos los derechos reservados.\n\n"
        "Herramienta de apoyo técnico en análisis de lubricación.\n\n"
        "Mobil™ es marca registrada de Exxon Mobil Corporation. "
        "Este software no representa afiliación oficial."
    )

# ─────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_excel(file) -> pd.DataFrame:
    return pd.read_excel(file)

def convert_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str).str.replace(r"[^0-9\.\-]", "", regex=True),
        errors="coerce"
    )

def first_non_null(series: pd.Series):
    s = series.dropna()
    return s.iloc[0] if not s.empty else None

def build_var_to_status(df: pd.DataFrame) -> dict:
    mapping = {}
    cols = set(df.columns)
    for c in df.columns:
        if isinstance(c, str) and " - Estado" in c:
            base = c.replace(" - Estado ", "").replace(" - Estado", "").strip()
            if base in cols:
                mapping[base] = c
    return mapping

def normalize_text(s: pd.Series) -> pd.Series:
    return (
        s.astype(str).str.strip().str.upper()
         .str.replace("Ó", "O").str.replace("Á", "A")
         .str.replace("É", "E").str.replace("Í", "I").str.replace("Ú", "U")
    )

def estado_clasificado(estado_raw: pd.Series) -> pd.Series:
    e = normalize_text(estado_raw)
    out = pd.Series(index=e.index, dtype="object")
    out[e.str.contains("ALERT",   na=False)] = "ALERTA"
    out[e.str.contains("ALERTA",  na=False)] = "ALERTA"
    out[e.str.contains("CAUTION", na=False)] = "PRECAUCION"
    out[e.str.contains("PRECAUC", na=False)] = "PRECAUCION"
    return out.fillna("NORMAL")

def apply_estado_filter(
    df_in: pd.DataFrame, var: str, excluir: bool, var_to_status: dict
) -> pd.Series:
    s = df_in[var]
    if not excluir:
        return s
    status_col = var_to_status.get(var)
    if not status_col or status_col not in df_in.columns:
        return s
    return s[estado_clasificado(df_in[status_col]) == "NORMAL"]

def clean_iqr(s: pd.Series) -> pd.Series:
    s = s.dropna()
    if len(s) < 4:
        return s
    q1, q3 = s.quantile(0.25), s.quantile(0.75)
    iqr = q3 - q1
    return s[(s >= q1 - 1.5 * iqr) & (s <= q3 + 1.5 * iqr)]

def calc_limits(series: pd.Series) -> dict:
    s = series.dropna()
    n_orig = int(len(s))

    base = {"n": n_orig, "n_orig": n_orig, "prec": np.nan, "alert": np.nan,
            "mean": np.nan, "std": np.nan, "median": np.nan,
            "vmin": np.nan, "vmax": np.nan, "confiabilidad": "⚠️ Insuficiente"}

    if n_orig < min_n:
        return {**base, "metodo": "Insuficiente"}

    if usar_iqr:
        s = clean_iqr(s)

    n = int(len(s))
    if n < min_n:
        return {**base, "n": n, "metodo": "Insuficiente (post-IQR)"}

    mean   = float(s.mean())
    std    = float(s.std(ddof=1)) if n > 1 else 0.0
    median = float(s.median())
    vmin   = float(s.min())
    vmax   = float(s.max())

    if n >= n_switch:
        prec  = float(s.quantile(p_prec  / 100))
        alert = float(s.quantile(p_alert / 100))
        metodo = f"Percentiles P{p_prec}/P{p_alert}"
    else:
        prec  = mean + k_prec  * std
        alert = mean + k_alert * std
        metodo = f"Media+{k_prec}σ / +{k_alert}σ"

    if n >= 30:
        confiabilidad = "★★★ Alta"
    elif n >= 10:
        confiabilidad = "★★☆ Media"
    else:
        confiabilidad = "★☆☆ Baja"

    return {"n": n, "n_orig": n_orig, "metodo": metodo,
            "prec": prec, "alert": alert,
            "mean": mean, "std": std, "median": median,
            "vmin": vmin, "vmax": vmax, "confiabilidad": confiabilidad}

def get_category(var: str) -> str:
    for cat, lst in vars_by_cat.items():
        if var in lst:
            return cat
    return "Sin categoría"

# ─────────────────────────────────────────────────────────────
# Excel con formato (openpyxl)
# ─────────────────────────────────────────────────────────────
def to_excel_colored(df_export: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Límites")
        ws = writer.sheets["Límites"]

        hdr_fill   = PatternFill("solid", fgColor="1A3A5C")
        prec_fill  = PatternFill("solid", fgColor="FEF9C3")
        alert_fill = PatternFill("solid", fgColor="FEE2E2")
        alt_fill   = PatternFill("solid", fgColor="F8FAFC")

        thin   = Side(border_style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        body_font = Font(size=10)

        cols_lower = [str(c).lower() for c in df_export.columns]
        prec_idx  = next((i + 1 for i, c in enumerate(cols_lower) if "precau" in c), None)
        alert_idx = next((i + 1 for i, c in enumerate(cols_lower) if "condenator" in c), None)

        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[1].height = 32

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_bg = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, cell in enumerate(row, start=1):
                cell.font      = body_font
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
                if col_idx == prec_idx:
                    cell.fill = prec_fill
                elif col_idx == alert_idx:
                    cell.fill = alert_fill
                else:
                    cell.fill = row_bg

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 38)

        ws.freeze_panes = "A2"

    return output.getvalue()

# ─────────────────────────────────────────────────────────────
# Función de estilo para la tabla de resultados
# ─────────────────────────────────────────────────────────────
def style_results(df_vis: pd.DataFrame) -> pd.DataFrame:
    style_df = pd.DataFrame("", index=df_vis.index, columns=df_vis.columns)

    if "Método" in df_vis.columns:
        mask_insuf = df_vis["Método"].astype(str).str.contains("Insuf", na=False)
        style_df[mask_insuf] = "background-color: #f3f4f6; color: #9ca3af"

    if "Límite de precaución" in df_vis.columns:
        mask = pd.to_numeric(df_vis["Límite de precaución"], errors="coerce").notna()
        style_df.loc[mask, "Límite de precaución"] = "background-color: #fef9c3; font-weight: bold"

    if "Límite condenatorio" in df_vis.columns:
        mask = pd.to_numeric(df_vis["Límite condenatorio"], errors="coerce").notna()
        style_df.loc[mask, "Límite condenatorio"] = "background-color: #fee2e2; font-weight: bold"

    return style_df

# ─────────────────────────────────────────────────────────────
# Carga del archivo
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">📂 Carga del archivo</div>', unsafe_allow_html=True)

col_up, col_info = st.columns([2, 1])
with col_up:
    archivos = st.file_uploader(
        "Selecciona uno o más Excel exportados desde SmartAssistance (ARCHIVO 2)",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Puedes seleccionar varios archivos a la vez. No se modifican, solo se leen en memoria."
    )
with col_info:
    st.info(
        "**Formato esperado:** Export estándar SmartAssistance\n\n"
        "**Columnas requeridas:** COMPONENTE, FECHA_INFORME\n\n"
        "**Columnas opcionales:** NOMBRE_OPERACION, TIPO_EQUIPO, PRODUCTO"
    )

if not archivos:
    st.markdown(
        "<div style='text-align:center; padding:3rem 0; color:#9ca3af;'>"
        "<div style='font-size:3rem;'>📊</div>"
        "<div style='font-size:1.1rem; margin-top:0.5rem;'>"
        "Carga uno o más archivos Excel para comenzar el análisis"
        "</div></div>",
        unsafe_allow_html=True
    )
    st.stop()

dfs_cargados = []
archivos_con_error = []
for arch in archivos:
    try:
        df_i = load_excel(arch).copy()
        df_i["ARCHIVO_ORIGEN"] = arch.name
        dfs_cargados.append(df_i)
    except Exception as e:
        archivos_con_error.append(f"{arch.name}: {e}")

if archivos_con_error:
    st.error("No se pudieron leer los siguientes archivos:\n" + "\n".join(archivos_con_error))

if not dfs_cargados:
    st.stop()

df = pd.concat(dfs_cargados, ignore_index=True)

if len(archivos) > 1:
    filas_por_archivo = {arch.name: len(df_i) for arch, df_i in zip(archivos, dfs_cargados)}
    resumen_archivos = " · ".join(f"**{n}**: {r:,} filas" for n, r in filas_por_archivo.items())
    st.success(f"{len(archivos)} archivos cargados y combinados — {len(df):,} registros totales: {resumen_archivos}")

required_cols = ["COMPONENTE", "FECHA_INFORME"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    st.error(f"**Faltan columnas requeridas:** {', '.join(missing)}")
    st.stop()

df["FECHA_INFORME"] = pd.to_datetime(df["FECHA_INFORME"], errors="coerce")
if df["FECHA_INFORME"].isna().all():
    st.error("La columna **FECHA_INFORME** no contiene fechas válidas.")
    st.stop()

col_operacion   = "NOMBRE_OPERACION" if "NOMBRE_OPERACION" in df.columns else None
col_tipo_equipo = "TIPO_EQUIPO"      if "TIPO_EQUIPO"      in df.columns else None
col_lubricante  = "PRODUCTO"         if "PRODUCTO"         in df.columns else None

var_to_status = build_var_to_status(df)

available   = set(df.columns)
vars_by_cat = {cat: [v for v in vlist if v in available] for cat, vlist in GUIDE_VARS.items()}
total_exist = sum(len(v) for v in vars_by_cat.values())

if total_exist == 0:
    st.error(
        "No se encontró ninguna variable de la guía en el archivo. "
        "Verifica que sea el export estándar de SmartAssistance (ARCHIVO 2)."
    )
    st.stop()

# ─────────────────────────────────────────────────────────────
# KPIs globales
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">📊 Resumen del archivo cargado</div>', unsafe_allow_html=True)

fecha_min_g  = df["FECHA_INFORME"].min()
fecha_max_g  = df["FECHA_INFORME"].max()

k1, k2, k3, k4, k5 = st.columns(5)
kpis = [
    (f"{len(df):,}",           "Total muestras"),
    (str(df["COMPONENTE"].nunique()), "Componentes"),
    (str(total_exist),         "Variables disponibles"),
    (fecha_min_g.strftime("%b %Y") if pd.notna(fecha_min_g) else "—", "Primera muestra"),
    (fecha_max_g.strftime("%b %Y") if pd.notna(fecha_max_g) else "—", "Última muestra"),
]
for col, (val, label) in zip([k1, k2, k3, k4, k5], kpis):
    with col:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-value">{val}</div>'
            f'<div class="kpi-label">{label}</div></div>',
            unsafe_allow_html=True
        )

with st.expander("🔍 Detalle de variables encontradas", expanded=False):
    for cat, icon in CAT_ICONS.items():
        found = vars_by_cat.get(cat, [])
        total = len(GUIDE_VARS.get(cat, []))
        st.write(f"**{icon} {cat}** — {len(found)} / {total} encontradas")
        miss = [v for v in GUIDE_VARS.get(cat, []) if v not in available]
        if miss:
            st.caption("No encontradas: " + " · ".join(miss))

# ─────────────────────────────────────────────────────────────
# 1. Filtros de análisis
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">1️⃣ Filtros de análisis</div>', unsafe_allow_html=True)
st.caption(
    "**Opcional.** Si quieres que los límites reflejen solo un contexto específico "
    "(por ejemplo, una mina, un tipo de equipo o un lubricante particular), "
    "activa los filtros que correspondan. Si no filtras, el cálculo usa todo el historial cargado."
)

df_f = df.copy()

usar_fechas = st.checkbox("Filtrar por rango de fechas", value=False)
if usar_fechas:
    fmin = df_f["FECHA_INFORME"].min()
    fmax = df_f["FECHA_INFORME"].max()
    if pd.notna(fmin) and pd.notna(fmax):
        ini, fin = st.date_input(
            "Rango de fechas",
            value=[fmin.date(), fmax.date()],
            min_value=fmin.date(),
            max_value=fmax.date()
        )
        df_f = df_f[
            (df_f["FECHA_INFORME"] >= pd.to_datetime(ini)) &
            (df_f["FECHA_INFORME"] <= pd.to_datetime(fin))
        ].copy()

c1, c2, c3 = st.columns(3)
with c1:
    if col_operacion:
        ops = st.multiselect("Operación", sorted(df_f[col_operacion].dropna().unique()))
        if ops:
            df_f = df_f[df_f[col_operacion].isin(ops)].copy()
    else:
        st.caption("NOMBRE_OPERACION no está en el archivo.")
with c2:
    if col_tipo_equipo:
        tipos = st.multiselect("Tipo de equipo", sorted(df_f[col_tipo_equipo].dropna().unique()))
        if tipos:
            df_f = df_f[df_f[col_tipo_equipo].isin(tipos)].copy()
    else:
        st.caption("TIPO_EQUIPO no está en el archivo.")
with c3:
    if col_lubricante:
        lubs = st.multiselect("Lubricante (PRODUCTO)", sorted(df_f[col_lubricante].dropna().unique()))
        if lubs:
            df_f = df_f[df_f[col_lubricante].isin(lubs)].copy()
    else:
        st.caption("PRODUCTO no está en el archivo.")

if df_f.empty:
    st.warning("No hay datos con los filtros seleccionados.")
    st.stop()

st.caption(f"Registros tras filtros: **{len(df_f):,}** de {len(df):,}")

# ─────────────────────────────────────────────────────────────
# 2. Inventario de componentes
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">2️⃣ Inventario de componentes</div>', unsafe_allow_html=True)
st.caption(
    "Esta tabla muestra cuántos análisis tiene cada componente en el historial. "
    "**Más muestras = límites más confiables.** "
    "Si un componente tiene menos de 10 muestras, los límites serán orientativos; "
    "con 30 o más, son estadísticamente robustos. "
    "Úsala para decidir cuáles vale la pena analizar."
)

group_cols = ["COMPONENTE"]
if col_operacion:   group_cols.append(col_operacion)
if col_tipo_equipo: group_cols.append(col_tipo_equipo)
if col_lubricante:  group_cols.append(col_lubricante)

inventario = (
    df_f.groupby(group_cols, dropna=False)
        .agg(
            Muestras      = ("COMPONENTE", "size"),
            Primera_fecha = ("FECHA_INFORME", "min"),
            Última_fecha  = ("FECHA_INFORME", "max"),
        )
        .reset_index()
        .sort_values("Muestras", ascending=False)
)
st.dataframe(inventario, use_container_width=True, height=260)

# ─────────────────────────────────────────────────────────────
# 3. Modo de cálculo
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">3️⃣ Modo de cálculo</div>', unsafe_allow_html=True)

with st.expander("❓ ¿Cuál modo debo elegir?", expanded=False):
    st.markdown("""
    **Límites por componente** *(opción recomendada cuando tienes suficiente historial)*
    Cada componente recibe sus propios límites, calculados únicamente con sus muestras.
    Ejemplo: el Motor 1 puede tener un límite de hierro distinto al Motor 2 si su operación
    o desgaste típico es diferente.

    **Límite único mezclando componentes** *(útil cuando cada componente tiene pocos datos)*
    Combina el historial de varios componentes similares para calcular un solo límite compartido.
    Ejemplo: si tienes 5 motores iguales con 8 muestras cada uno, mezclarlos da 40 muestras
    y un límite mucho más confiable. Úsalo cuando los equipos sean del mismo modelo y operación.
    """)

modo = st.radio(
    "¿Cómo quieres calcular los límites?",
    options=["Límites por componente", "Límite único mezclando varios componentes"],
    horizontal=True,
    index=0
)

# ─────────────────────────────────────────────────────────────
# 4. Selección de componentes
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">4️⃣ Selección de componentes</div>', unsafe_allow_html=True)

componentes = sorted(df_f["COMPONENTE"].dropna().astype(str).unique())
if not componentes:
    st.warning("No hay componentes disponibles con los filtros actuales.")
    st.stop()

if modo == "Límites por componente":
    comps_sel = st.multiselect(
        "Componentes a analizar",
        options=componentes,
        default=componentes[:1],
        help="Se calculan límites independientes para cada uno."
    )
    if not comps_sel:
        st.warning("Selecciona al menos un componente.")
        st.stop()
    df_calc         = df_f[df_f["COMPONENTE"].astype(str).isin(set(map(str, comps_sel)))].copy()
    etiqueta_mezcla = None
else:
    n_mix = st.number_input(
        "Cantidad de componentes a mezclar",
        min_value=2, max_value=len(componentes), value=min(2, len(componentes)), step=1
    )
    comps_mix = st.multiselect(
        "Componentes a mezclar",
        options=componentes, default=componentes[:n_mix], max_selections=n_mix
    )
    if len(comps_mix) != n_mix:
        st.warning(f"Selecciona exactamente {n_mix} componentes.")
        st.stop()
    df_calc         = df_f[df_f["COMPONENTE"].astype(str).isin(set(map(str, comps_mix)))].copy()
    etiqueta_mezcla = " + ".join(comps_mix)
    st.info(f"Límite único para: **{etiqueta_mezcla}** — {len(df_calc):,} registros combinados")

if df_calc.empty:
    st.warning("No hay registros para la selección actual.")
    st.stop()

# ─────────────────────────────────────────────────────────────
# 5. Variables para el cálculo
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">5️⃣ Variables para el cálculo</div>', unsafe_allow_html=True)

with st.expander("❓ ¿Qué variable debo seleccionar?", expanded=False):
    st.markdown("""
    Las variables están agrupadas por categoría. Selecciona las que quieres incluir en el análisis:

    | Categoría | Qué mide | Ejemplos de uso |
    |-----------|----------|-----------------|
    | ⚙️ **Desgaste** | Metales que se desprenden de las piezas en movimiento | Hierro (cuerpo/cilindros), Cobre (cojinetes/enfriadores), PQI (partículas grandes) |
    | 🧪 **Propiedades del lubricante** | Condición del aceite en sí | Viscosidad (fluidez), BN (reserva alcalina), Oxidación (degradación del aceite) |
    | ⚠️ **Contaminantes** | Sustancias que no deberían estar en el aceite | Agua (fuga de refrigerante o condensación), Silicio (polvo/tierra), Sodio (refrigerante) |
    | 🔬 **Aditivos** | Paquete de aditivos del lubricante | Calcio, Zinc, Fósforo (se consumen con el uso; su baja indica que el aceite está agotado) |

    **Recomendación:** selecciona todas las variables disponibles para tu componente.
    Puedes calcular límites para todas de una vez y solo exportar las que uses en tus reportes.
    """)

excluir_fuera_normal = st.toggle(
    "Excluir resultados fuera de lo normal (Precaución y Alerta) del cálculo base",
    value=True,
    help=(
        "ACTIVADO (recomendado): la app calcula los límites usando solo las muestras 'Normales', "
        "ignorando las que ya estaban en Precaución o Alerta. "
        "Esto evita que valores de falla pasada inflen el límite y lo hagan demasiado permisivo. "
        "DESACTIVADO: usa todas las muestras, incluidas las que tuvieron algún problema."
    )
)

if "vars_checked" not in st.session_state:
    st.session_state["vars_checked"] = set()

def render_cat(title: str, vars_list: list, expanded: bool):
    if not vars_list:
        return
    icon = CAT_ICONS.get(title, "")
    with st.expander(f"{icon} {title}  ({len(vars_list)} variables disponibles)", expanded=expanded):
        cols = st.columns(3)
        for i, v in enumerate(vars_list):
            with cols[i % 3]:
                unit  = VAR_UNITS.get(v, "")
                label = f"{v}" + (f"  *[{unit}]*" if unit else "")
                cur   = v in st.session_state["vars_checked"]
                val   = st.checkbox(label, value=cur, key=f"chk_{title}_{v}")
                if val:
                    st.session_state["vars_checked"].add(v)
                else:
                    st.session_state["vars_checked"].discard(v)

render_cat("Desgaste",                   vars_by_cat["Desgaste"],                   expanded=True)
render_cat("Propiedades del lubricante", vars_by_cat["Propiedades del lubricante"], expanded=True)
render_cat("Contaminantes",              vars_by_cat["Contaminantes"],              expanded=True)
render_cat("Aditivos",                   vars_by_cat["Aditivos"],                   expanded=False)

vars_sel = sorted(list(st.session_state["vars_checked"]))
if not vars_sel:
    st.warning("Selecciona al menos una variable.")
    st.stop()

st.success(f"**{len(vars_sel)}** variable(s) seleccionada(s): {', '.join(vars_sel)}")

for v in vars_sel:
    df_calc[v] = convert_numeric(df_calc[v])

# ─────────────────────────────────────────────────────────────
# 6. Calcular y mostrar resultados
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-title">6️⃣ Resultados y descarga</div>', unsafe_allow_html=True)
st.caption(
    "Al presionar el botón, la app calcula los límites y muestra: "
    "la tabla de resultados con código de color, "
    "el resumen de cuántos límites se pudieron calcular, "
    "y los histogramas de distribución por variable. "
    "Luego puedes descargar el archivo en Excel (con formato de colores) o CSV para Power BI."
)

if st.button("🔢 Calcular límites", type="primary"):
    filas = []

    with st.spinner("Calculando límites..."):
        if modo == "Límites por componente":
            for comp, g in df_calc.groupby(df_calc["COMPONENTE"].astype(str)):
                for var in vars_sel:
                    serie = apply_estado_filter(g, var, excluir_fuera_normal, var_to_status)
                    out   = calc_limits(serie)
                    filas.append({
                        "Operación":             first_non_null(g[col_operacion])   if col_operacion   else None,
                        "Tipo de equipo":        first_non_null(g[col_tipo_equipo]) if col_tipo_equipo else None,
                        "Lubricante":            first_non_null(g[col_lubricante])  if col_lubricante  else None,
                        "Componente":            comp,
                        "Categoría":             get_category(var),
                        "Variable":              var,
                        "Unidad":                VAR_UNITS.get(var, ""),
                        "Datos excluidos":       "Sí" if excluir_fuera_normal else "No",
                        "IQR aplicado":          "Sí" if usar_iqr else "No",
                        "Datos válidos (n)":     out["n"],
                        "n original":            out.get("n_orig", out["n"]),
                        "Método":                out["metodo"],
                        "Mínimo":                out["vmin"],
                        "Máximo":                out["vmax"],
                        "Promedio":              out["mean"],
                        "Desviación estándar":   out["std"],
                        "Mediana":               out["median"],
                        "Límite de precaución":  out["prec"],
                        "Límite condenatorio":   out["alert"],
                        "Confiabilidad":         out.get("confiabilidad", ""),
                        "Primera fecha":         g["FECHA_INFORME"].min(),
                        "Última fecha":          g["FECHA_INFORME"].max(),
                    })
        else:
            for var in vars_sel:
                serie = apply_estado_filter(df_calc, var, excluir_fuera_normal, var_to_status)
                out   = calc_limits(serie)
                filas.append({
                    "Componentes mezclados": etiqueta_mezcla,
                    "Categoría":             get_category(var),
                    "Variable":              var,
                    "Unidad":                VAR_UNITS.get(var, ""),
                    "Datos excluidos":       "Sí" if excluir_fuera_normal else "No",
                    "IQR aplicado":          "Sí" if usar_iqr else "No",
                    "Datos válidos (n)":     out["n"],
                    "n original":            out.get("n_orig", out["n"]),
                    "Método":                out["metodo"],
                    "Mínimo":                out["vmin"],
                    "Máximo":                out["vmax"],
                    "Promedio":              out["mean"],
                    "Desviación estándar":   out["std"],
                    "Mediana":               out["median"],
                    "Límite de precaución":  out["prec"],
                    "Límite condenatorio":   out["alert"],
                    "Confiabilidad":         out.get("confiabilidad", ""),
                    "Primera fecha":         df_calc["FECHA_INFORME"].min(),
                    "Última fecha":          df_calc["FECHA_INFORME"].max(),
                })

    resultados = pd.DataFrame(filas)

    # ── Redondeo para visualización
    dec = st.number_input("Decimales para visualización", min_value=0, value=1, step=1)
    vista = resultados.copy()
    num_cols_vis = [
        "Mínimo", "Máximo", "Promedio", "Desviación estándar", "Mediana",
        "Límite de precaución", "Límite condenatorio"
    ]
    for c in num_cols_vis:
        if c in vista.columns:
            vista[c] = pd.to_numeric(vista[c], errors="coerce").round(dec)

    # ── Orden de columnas
    if modo == "Límites por componente":
        orden = [
            "Operación", "Tipo de equipo", "Lubricante", "Componente",
            "Categoría", "Variable", "Unidad",
            "Datos excluidos", "IQR aplicado",
            "Datos válidos (n)", "n original", "Método", "Confiabilidad",
            "Mínimo", "Máximo", "Promedio", "Desviación estándar", "Mediana",
            "Límite de precaución", "Límite condenatorio",
            "Primera fecha", "Última fecha",
        ]
    else:
        orden = [
            "Componentes mezclados", "Categoría", "Variable", "Unidad",
            "Datos excluidos", "IQR aplicado",
            "Datos válidos (n)", "n original", "Método", "Confiabilidad",
            "Mínimo", "Máximo", "Promedio", "Desviación estándar", "Mediana",
            "Límite de precaución", "Límite condenatorio",
            "Primera fecha", "Última fecha",
        ]

    columnas = [c for c in orden if c in vista.columns]
    styled   = vista[columnas].style.apply(style_results, axis=None)

    with st.expander("📋 ¿Qué significa cada columna?", expanded=False):
        st.markdown("""
        | Columna | Significado |
        |---------|-------------|
        | **Datos válidos (n)** | Cantidad de análisis usados para calcular (después de filtros y exclusiones). |
        | **n original** | Total de análisis antes de limpiar outliers o excluir estados anormales. |
        | **Método** | Cómo se calcularon los límites: *Percentiles* si había muchos datos, *Media+kσ* si había pocos. |
        | **Mínimo / Máximo** | El valor más bajo y más alto que se ha visto históricamente para ese componente y variable. |
        | **Promedio** | El valor típico del historial. |
        | **Desviación estándar** | Qué tanto varían los valores. Alta variación → mayor incertidumbre en los límites. |
        | **Mediana** | El valor central del historial (menos afectado por valores extremos que el promedio). |
        | 🟡 **Límite de Precaución** | Umbral de vigilancia. Si un análisis supera este valor, el componente merece atención. |
        | 🔴 **Límite Condenatorio** | Umbral de acción. Si supera este valor, hay alta probabilidad de falla activa. |
        | **Primera / Última fecha** | Rango de fechas del historial usado en el cálculo. |

        **Colores en la tabla:**
        - 🟡 Amarillo = columna Precaución (umbral de vigilancia)
        - 🔴 Rojo claro = columna Condenatorio (umbral de acción)
        - ⬜ Gris = fila con datos insuficientes (no se calcularon límites)
        """)

    st.dataframe(styled, use_container_width=True, height=460)

    # ── Métricas de resumen
    mask_insuf   = resultados["Método"].astype(str).str.contains("Insuf", na=False)
    total_calc_r = resultados[~mask_insuf]
    total_insuf  = resultados[mask_insuf]

    m1, m2, m3 = st.columns(3)
    m1.metric("Límites calculados",         len(total_calc_r))
    m2.metric("Sin suficientes datos",      len(total_insuf))
    if modo == "Límites por componente" and "Componente" in resultados.columns:
        m3.metric("Componentes procesados", resultados["Componente"].nunique())
    else:
        m3.metric("Variables procesadas",   len(vars_sel))

    # ── Semáforo: estado actual vs límites calculados
    if modo == "Límites por componente" and "Componente" in total_calc_r.columns:
        st.markdown('<div class="section-title">🚦 Estado actual vs límites calculados</div>', unsafe_allow_html=True)
        st.caption(
            "Cruza el **último análisis real** de cada componente contra los límites recién calculados. "
            "Así sabes qué equipos ya superan hoy sus propios límites históricos."
        )

        sem_filas = []
        for _, lim_row in total_calc_r.iterrows():
            comp      = lim_row.get("Componente")
            var       = lim_row.get("Variable")
            prec_lim  = lim_row.get("Límite de precaución")
            alert_lim = lim_row.get("Límite condenatorio")
            unit      = lim_row.get("Unidad", "")

            df_comp = df_calc[df_calc["COMPONENTE"].astype(str) == str(comp)]
            df_comp_sorted = df_comp.sort_values("FECHA_INFORME", ascending=False)

            val, fecha_ult = np.nan, None
            if not df_comp_sorted.empty and var in df_comp_sorted.columns:
                row_ult  = df_comp_sorted.iloc[0]
                val      = row_ult[var]
                fecha_ult = row_ult.get("FECHA_INFORME")

            try:
                val_f = float(val)
            except (TypeError, ValueError):
                val_f = np.nan

            if pd.isna(val_f):
                estado_actual = "⬜ Sin dato"
                pct_txt = "—"
            elif pd.notna(alert_lim) and val_f >= float(alert_lim):
                estado_actual = "🔴 Supera Condenatorio"
                pct = round(val_f / float(alert_lim) * 100) if float(alert_lim) > 0 else None
                pct_txt = f"{pct}% del lím. cond." if pct else "—"
            elif pd.notna(prec_lim) and val_f >= float(prec_lim):
                estado_actual = "🟡 Supera Precaución"
                pct = round(val_f / float(prec_lim) * 100) if float(prec_lim) > 0 else None
                pct_txt = f"{pct}% del lím. prec." if pct else "—"
            elif pd.notna(prec_lim):
                estado_actual = "🟢 Normal"
                pct = round(val_f / float(prec_lim) * 100) if float(prec_lim) > 0 else None
                pct_txt = f"{pct}% del lím. prec." if pct else "—"
            else:
                estado_actual = "⬜ Sin límite"
                pct_txt = "—"

            sem_filas.append({
                "Componente":       comp,
                "Variable":         var,
                "Unidad":           unit,
                "Última muestra":   str(fecha_ult)[:10] if fecha_ult else "—",
                "Valor actual":     round(val_f, dec) if pd.notna(val_f) else "—",
                "Lím. Precaución":  round(float(prec_lim), dec) if pd.notna(prec_lim) else "—",
                "Lím. Condenatorio":round(float(alert_lim), dec) if pd.notna(alert_lim) else "—",
                "% respecto al límite": pct_txt,
                "Estado actual":    estado_actual,
            })

        df_sem = pd.DataFrame(sem_filas)

        def style_semaforo(df_s: pd.DataFrame) -> pd.DataFrame:
            style = pd.DataFrame("", index=df_s.index, columns=df_s.columns)
            if "Estado actual" not in df_s.columns:
                return style
            for idx, row in df_s.iterrows():
                est = str(row.get("Estado actual", ""))
                if "Condenatorio" in est:
                    style.loc[idx]               = "background-color: #fee2e2"
                    style.loc[idx, "Estado actual"] = "background-color: #fee2e2; font-weight:bold; color:#dc2626"
                elif "Precaución" in est:
                    style.loc[idx]               = "background-color: #fef9c3"
                    style.loc[idx, "Estado actual"] = "background-color: #fef9c3; font-weight:bold; color:#d97706"
                elif "Normal" in est:
                    style.loc[idx, "Estado actual"] = "color:#16a34a; font-weight:bold"
            return style

        st.dataframe(df_sem.style.apply(style_semaforo, axis=None), use_container_width=True, height=380)

        n_ok   = sum(1 for r in sem_filas if "Normal"       in r["Estado actual"])
        n_prec = sum(1 for r in sem_filas if "Precaución"   in r["Estado actual"])
        n_cond = sum(1 for r in sem_filas if "Condenatorio" in r["Estado actual"])
        sa1, sa2, sa3 = st.columns(3)
        sa1.metric("🟢 Normales",         n_ok)
        sa2.metric("🟡 En Precaución",    n_prec)
        sa3.metric("🔴 En Condenatorio",  n_cond)

    # ── Gráficas interactivas por variable (Plotly)
    max_vars_plot = 12
    if not total_calc_r.empty and len(vars_sel) <= max_vars_plot:
        st.markdown('<div class="section-title">📈 Análisis gráfico por variable</div>', unsafe_allow_html=True)
        st.caption("Selecciona una variable para ver su distribución histórica y su evolución en el tiempo.")

        var_opciones = [v for v in vars_sel if not total_calc_r[total_calc_r["Variable"] == v].empty]
        var_elegida  = st.selectbox("Variable a visualizar", options=var_opciones,
                                    format_func=lambda v: f"{v}  [{VAR_UNITS.get(v,'')}]" if VAR_UNITS.get(v) else v)

        if var_elegida:
            subset_g = total_calc_r[total_calc_r["Variable"] == var_elegida]
            unit_g   = VAR_UNITS.get(var_elegida, "")

            if modo == "Límites por componente" and "Componente" in subset_g.columns:
                comp_list_g = subset_g["Componente"].unique().tolist()
            else:
                comp_list_g = [etiqueta_mezcla or "Mezcla"]

            comp_ver = st.selectbox("Componente", options=comp_list_g) if len(comp_list_g) > 1 else comp_list_g[0]

            if modo == "Límites por componente" and "Componente" in subset_g.columns:
                data_hist = df_calc[df_calc["COMPONENTE"].astype(str) == str(comp_ver)]
                row_lim_g = subset_g[subset_g["Componente"] == comp_ver]
            else:
                data_hist = df_calc.copy()
                row_lim_g = subset_g

            row_lim_g = row_lim_g.iloc[0] if not row_lim_g.empty else pd.Series()
            p_val_g   = row_lim_g.get("Límite de precaución", np.nan)
            a_val_g   = row_lim_g.get("Límite condenatorio",  np.nan)
            mean_g    = row_lim_g.get("Promedio", np.nan)

            ts_data = (
                data_hist[["FECHA_INFORME", var_elegida]]
                .dropna()
                .sort_values("FECHA_INFORME")
                .copy()
            )
            ts_data[var_elegida] = pd.to_numeric(ts_data[var_elegida], errors="coerce")
            ts_data = ts_data.dropna()

            tab_ts, tab_dist = st.tabs(["📈 Tendencia en el tiempo", "📊 Distribución histórica"])

            # ── Tab 1: Serie de tiempo ──────────────────────────────
            with tab_ts:
                if ts_data.empty:
                    st.info("No hay datos históricos suficientes para este componente y variable.")
                else:
                    colors_ts = []
                    for val_ts in ts_data[var_elegida]:
                        if pd.notna(a_val_g) and val_ts >= float(a_val_g):
                            colors_ts.append("#dc2626")
                        elif pd.notna(p_val_g) and val_ts >= float(p_val_g):
                            colors_ts.append("#d97706")
                        else:
                            colors_ts.append("#3b82f6")

                    fig_ts = go.Figure()

                    # Zonas de color de fondo
                    y_max = ts_data[var_elegida].max()
                    if pd.notna(p_val_g) and pd.notna(a_val_g):
                        fig_ts.add_hrect(y0=float(p_val_g), y1=float(a_val_g),
                                         fillcolor="#fef3c7", opacity=0.35, line_width=0,
                                         annotation_text="Zona Precaución", annotation_position="top left",
                                         annotation_font_color="#d97706", annotation_font_size=10)
                        fig_ts.add_hrect(y0=float(a_val_g), y1=max(y_max * 1.15, float(a_val_g) * 1.1),
                                         fillcolor="#fee2e2", opacity=0.35, line_width=0,
                                         annotation_text="Zona Condenatorio", annotation_position="top left",
                                         annotation_font_color="#dc2626", annotation_font_size=10)

                    # Línea de tendencia
                    fig_ts.add_trace(go.Scatter(
                        x=ts_data["FECHA_INFORME"],
                        y=ts_data[var_elegida],
                        mode="lines+markers",
                        name=var_elegida,
                        line=dict(color="#94a3b8", width=2),
                        marker=dict(color=colors_ts, size=10, line=dict(color="white", width=1.5)),
                        hovertemplate=(
                            "<b>%{x|%d %b %Y}</b><br>"
                            + (f"{var_elegida}: " if len(var_elegida) < 25 else "Valor: ")
                            + "%{y:.2f} " + unit_g + "<extra></extra>"
                        ),
                    ))

                    # Líneas de límite
                    if pd.notna(p_val_g):
                        fig_ts.add_hline(y=float(p_val_g), line_dash="dash", line_color="#d97706", line_width=2,
                                         annotation_text=f"Precaución {float(p_val_g):.1f} {unit_g}",
                                         annotation_position="bottom right",
                                         annotation_font_color="#d97706")
                    if pd.notna(a_val_g):
                        fig_ts.add_hline(y=float(a_val_g), line_dash="solid", line_color="#dc2626", line_width=2,
                                         annotation_text=f"Condenatorio {float(a_val_g):.1f} {unit_g}",
                                         annotation_position="bottom right",
                                         annotation_font_color="#dc2626")
                    if pd.notna(mean_g):
                        fig_ts.add_hline(y=float(mean_g), line_dash="dot", line_color="#1a3a5c", line_width=1.5,
                                         annotation_text=f"Media {float(mean_g):.1f}",
                                         annotation_position="top right",
                                         annotation_font_color="#1a3a5c")

                    fig_ts.update_layout(
                        height=380,
                        margin=dict(t=30, r=30, b=50, l=60),
                        plot_bgcolor="#fafafa",
                        paper_bgcolor="white",
                        font=dict(family="Segoe UI, Inter, sans-serif", size=12),
                        xaxis=dict(title="Fecha de informe", showgrid=True, gridcolor="#f0f0f0", tickangle=-30),
                        yaxis=dict(title=unit_g if unit_g else "Valor", showgrid=True, gridcolor="#f0f0f0"),
                        hovermode="x unified",
                        showlegend=False,
                    )
                    st.plotly_chart(fig_ts, use_container_width=True)
                    st.caption(
                        "🔵 Normal · 🟡 Supera Precaución · 🔴 Supera Condenatorio  "
                        "— cada punto es una muestra de análisis de aceite"
                    )

            # ── Tab 2: Distribución ─────────────────────────────────
            with tab_dist:
                if ts_data.empty:
                    st.info("No hay datos para mostrar.")
                else:
                    fig_dist = go.Figure()
                    fig_dist.add_trace(go.Histogram(
                        x=ts_data[var_elegida],
                        nbinsx=min(25, max(6, len(ts_data) // 3 + 1)),
                        marker_color="#3b82f6",
                        marker_line_color="white",
                        marker_line_width=0.8,
                        opacity=0.75,
                        name="Frecuencia",
                        hovertemplate="Valor: %{x:.1f}<br>Frecuencia: %{y}<extra></extra>",
                    ))
                    if pd.notna(p_val_g):
                        fig_dist.add_vline(x=float(p_val_g), line_dash="dash", line_color="#d97706", line_width=2.5,
                                           annotation_text=f"Precaución: {float(p_val_g):.1f}",
                                           annotation_position="top", annotation_font_color="#d97706")
                    if pd.notna(a_val_g):
                        fig_dist.add_vline(x=float(a_val_g), line_dash="solid", line_color="#dc2626", line_width=2.5,
                                           annotation_text=f"Condenatorio: {float(a_val_g):.1f}",
                                           annotation_position="top", annotation_font_color="#dc2626")
                    if pd.notna(mean_g):
                        fig_dist.add_vline(x=float(mean_g), line_dash="dot", line_color="#1a3a5c", line_width=2,
                                           annotation_text=f"Media: {float(mean_g):.1f}",
                                           annotation_position="top", annotation_font_color="#1a3a5c")

                    fig_dist.update_layout(
                        height=360,
                        margin=dict(t=30, r=30, b=50, l=60),
                        plot_bgcolor="#fafafa",
                        paper_bgcolor="white",
                        font=dict(family="Segoe UI, Inter, sans-serif", size=12),
                        xaxis=dict(title=unit_g if unit_g else "Valor", showgrid=True, gridcolor="#f0f0f0"),
                        yaxis=dict(title="Número de muestras", showgrid=True, gridcolor="#f0f0f0"),
                        showlegend=False,
                        bargap=0.05,
                    )
                    st.plotly_chart(fig_dist, use_container_width=True)
                    st.caption(
                        "Cada barra representa cuántas muestras tuvieron ese rango de valores. "
                        "La mayoría debería estar a la izquierda de la línea de Precaución."
                    )

    elif len(vars_sel) > max_vars_plot:
        st.info(f"Las gráficas se muestran cuando hay ≤ {max_vars_plot} variables seleccionadas.")

    # ── Descarga
    st.markdown("---")
    filename_base = (
        "limites_por_componente"
        if modo == "Límites por componente"
        else "limites_mezcla_componentes"
    )

    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "⬇️ Descargar Excel con formato",
            data=to_excel_colored(vista[columnas]),
            file_name=f"{filename_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with dl2:
        st.download_button(
            "⬇️ Descargar CSV (Power BI / Excel simple)",
            data=vista[columnas].to_csv(index=False).encode("utf-8"),
            file_name=f"{filename_base}.csv",
            mime="text/csv",
            use_container_width=True
        )

else:
    st.info(
        "⬆️ Completa los pasos anteriores y presiona **Calcular límites** para ver los resultados."
    )

# ─────────────────────────────────────────────────────────────
# Footer
# ─────────────────────────────────────────────────────────────
st.markdown(
    "<div class='footer'>"
    "© 2026 · Javier Parada · Análisis de Lubricación · "
    "Mobil™ es marca registrada de Exxon Mobil Corporation · Uso interno"
    "</div>",
    unsafe_allow_html=True
)
